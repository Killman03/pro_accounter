import pandas as pd
import re
from io import BytesIO
from openpyxl.utils import get_column_letter
import pandas as pd

def generate_excel_report(active_machines_data: list[dict], payments_data: list[dict], closed_machines_data: list[dict] = None) -> BytesIO:
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Первый лист - активные кофемашины
        df_active = pd.DataFrame(active_machines_data)
        df_active.to_excel(writer, sheet_name='Активные сделки', index=False)
        
        # Второй лист - платежи
        df_payments = pd.DataFrame(payments_data)
        df_payments.to_excel(writer, sheet_name='Платежи', index=False)
        
        # Третий лист - закрытые сделки (если есть данные)
        if closed_machines_data:
            df_closed = pd.DataFrame(closed_machines_data)
            df_closed.to_excel(writer, sheet_name='Закрытые сделки', index=False)

        _autosize_all_sheets(writer)
    
    output.seek(0)
    return output 


def generate_profit_share_report(rows: list[dict]) -> BytesIO:
    """
    Формирует Excel с распределением прибыли.
    Один лист = месяц старта арендатора, строки — все выплаты по арендатору.
    """
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df = pd.DataFrame(rows) if rows else pd.DataFrame([])
        wrote_any = False
        if df.empty:
            df.to_excel(writer, sheet_name="Нет данных", index=False)
            wrote_any = True
        else:
            if "Дата платежа" in df.columns:
                df.sort_values(by="Дата платежа", inplace=True, ignore_index=True)

            # убираем месячные столбцы без выплат (все пусто или 0)
            month_cols = [c for c in df.columns if re.fullmatch(r"\d{4}-\d{2}", str(c))]
            drop_cols = []
            for c in month_cols:
                col = df[c]
                if col.isna().all() or col.fillna(0).eq(0).all():
                    drop_cols.append(c)
            if drop_cols:
                df = df.drop(columns=drop_cols)

            if "Месяц старта" in df.columns:
                for month, df_month in df.groupby("Месяц старта"):
                    df_month.to_excel(writer, sheet_name=month, index=False)
                    wrote_any = True
            else:
                df.to_excel(writer, sheet_name="Выплаты", index=False)
                wrote_any = True
        # Гарантируем наличие видимого листа
        if not wrote_any or not writer.book.worksheets:
            ws = writer.book.create_sheet("Нет данных")
            ws.sheet_state = "visible"
        else:
            first = True
            for ws in writer.book.worksheets:
                ws.sheet_state = "visible"
                if first:
                    writer.book.active = writer.book.sheetnames.index(ws.title)
                    first = False
        _autosize_all_sheets(writer)
    output.seek(0)
    return output


def _autosize_all_sheets(writer: pd.ExcelWriter) -> None:
    """Устанавливает ширину столбцов по максимальной длине текста в каждом столбце."""
    wb = writer.book
    if not wb.worksheets:
        return
    for ws in wb.worksheets:
        for idx, column_cells in enumerate(ws.columns, start=1):
            max_length = 0
            for cell in column_cells:
                try:
                    cell_value = "" if cell.value is None else str(cell.value)
                    max_length = max(max_length, len(cell_value))
                except Exception:
                    continue
            adjusted_width = max_length + 2 if max_length else 10
            ws.column_dimensions[get_column_letter(idx)].width = adjusted_width